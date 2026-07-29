import re

import openpyxl
import pandas as pd
import streamlit as st

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(
    page_title="Social Investment Managers & Advisors LLC - CFO Dashboard",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Social Investment Managers & Advisors LLC")
st.markdown("### Executive Financial Performance & Monthly Expense Analytics")
st.markdown(
    "**Developed by: Abdul Rehman — VP Finance & CFO**  \n*CFO Data"
    " Analytics Workspace | Focus: Monthly P&L Trends, Major Expense Categories &"
    " Balance Sheet Position*"
)
st.divider()

MONTH_RE = re.compile(
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s*'?\s*(\d{2,4})",
    re.IGNORECASE,
)
NET_LABELS = {"net income", "net operating income", "net other income", "gross profit"}


# =============================================================================
# PARSER — tailored to QBO's "P&L by Month (with PY comparison)" and
# "Balance Sheet" xlsx exports. Both use openpyxl (not pandas) so we can read
# each row's real indentation, which is the only signal that distinguishes a
# section header from a line item from a subtotal.
# =============================================================================
class ParsedReport:
    def __init__(self):
        self.month_cols = []  # [{name, current_col, py_col}]
        self.total_current_col = None
        self.total_py_col = None
        self.rows = []  # [{label, indent, is_total, is_net, section, values: {col: val}}]

    def total_of(self, name):
        """Match a 'Total for <name>' row (QBO's subtotal label pattern)."""
        target = f"total for {name.lower()}"
        for r in self.rows:
            if r["label"].strip().lower() == target:
                return r
        return None

    def net_row(self, name):
        target = name.lower()
        for r in self.rows:
            if r["is_net"] and r["label"].strip().lower() == target:
                return r
        return None

    def month_current(self, row, month_name):
        for m in self.month_cols:
            if m["name"] == month_name:
                return row["values"].get(m["current_col"])
        return None

    def month_py(self, row, month_name):
        for m in self.month_cols:
            if m["name"] == month_name and m["py_col"]:
                return row["values"].get(m["py_col"])
        return None

    def current_total(self, row):
        if row is None or self.total_current_col is None:
            return None
        return row["values"].get(self.total_current_col)

    def py_total_full_year(self, row):
        if row is None or self.total_py_col is None:
            return None
        return row["values"].get(self.total_py_col)


def parse_qbo_report(file) -> ParsedReport:
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    report = ParsedReport()

    # --- locate header row: the row with an exact "Total" cell.
    # (We deliberately don't match on loose month-substrings for header
    # detection -- title rows like "As of Jul 28, 2026" contain "Jul" and
    # would false-positive.) ---
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=12):
        for cell in row:
            if cell.value is not None and str(cell.value).strip().lower() == "total":
                header_row_idx = row[0].row
                break
        if header_row_idx:
            break
    if header_row_idx is None:
        raise ValueError(
            "Could not find a 'Total' column header in the first 12 rows. "
            "This doesn't look like a standard QBO P&L or Balance Sheet export."
        )

    header_row = ws[header_row_idx]
    sub_row_idx = header_row_idx + 1
    sub_row = ws[sub_row_idx]
    has_py_pairs = any(
        isinstance(c.value, str) and c.value.strip().upper() == "CURRENT" for c in sub_row
    )

    col = 1
    max_col = ws.max_column
    while col <= max_col:
        cell = header_row[col - 1]
        val = cell.value
        if val and MONTH_RE.search(str(val)):
            month_name = MONTH_RE.search(str(val)).group(0)
            current_col = col
            py_col = None
            if has_py_pairs and col < max_col:
                next_val = sub_row[col].value
                if next_val and "(py)" in str(next_val).lower():
                    py_col = col + 1
            report.month_cols.append(
                {"name": month_name, "current_col": current_col, "py_col": py_col}
            )
            col += 2 if py_col else 1
        elif val and str(val).strip().lower() == "total":
            report.total_current_col = col
            if has_py_pairs and col < max_col:
                next_val = sub_row[col].value
                if next_val and "(py)" in str(next_val).lower():
                    report.total_py_col = col + 1
            col += 2 if report.total_py_col else 1
        else:
            col += 1

    data_start = sub_row_idx + 1 if has_py_pairs else header_row_idx + 1

    current_section = None
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        label_cell = row[0]
        label = label_cell.value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()

        indent = 0
        try:
            if label_cell.alignment and label_cell.alignment.indent:
                indent = int(label_cell.alignment.indent)
        except Exception:
            indent = 0

        is_total = label.lower().startswith("total for ") or label.lower().startswith("total ")
        is_net = label.lower() in NET_LABELS

        if indent == 0 and not is_total and not is_net:
            current_section = label

        values = {}
        for m in report.month_cols:
            values[m["current_col"]] = ws.cell(row=label_cell.row, column=m["current_col"]).value
            if m["py_col"]:
                values[m["py_col"]] = ws.cell(row=label_cell.row, column=m["py_col"]).value
        if report.total_current_col:
            values[report.total_current_col] = ws.cell(
                row=label_cell.row, column=report.total_current_col
            ).value
        if report.total_py_col:
            values[report.total_py_col] = ws.cell(
                row=label_cell.row, column=report.total_py_col
            ).value

        report.rows.append(
            {
                "label": label,
                "indent": indent,
                "is_total": is_total,
                "is_net": is_net,
                "section": current_section,
                "values": values,
            }
        )

    return report


def category_breakdown(report: ParsedReport, section_name: str):
    """
    Return [(category_name, ytd_current_amount)] for the indent-1 rows under a
    top-level section (e.g. 'Expenses'). QBO accounts can have sub-accounts
    (their own 'Total for X' row) AND/OR a direct value on the parent row
    itself -- when a 'Total for X' row exists for a category, it already
    includes the parent's direct postings, so we use it and skip the raw
    parent row to avoid double-counting.
    """
    indent1 = [r for r in report.rows if r["section"] == section_name and r["indent"] == 1]
    total_rows = {
        r["label"][len("total for "):].strip().lower(): r for r in indent1 if r["is_total"]
    }
    cats = []
    for r in indent1:
        if r["is_total"]:
            continue
        if r["label"].strip().lower() in total_rows:
            continue  # superseded by its own Total row
        val = report.current_total(r) or 0
        if val:
            cats.append((r["label"], val))
    for key, r in total_rows.items():
        cats.append((r["label"][len("Total for "):], report.current_total(r) or 0))
    return sorted(cats, key=lambda x: -x[1])


def ytd_cutoff_index(report: ParsedReport, income_row):
    if income_row is None or not report.month_cols:
        return len(report.month_cols) - 1
    vals = [report.month_current(income_row, m["name"]) or 0 for m in report.month_cols]
    nonzero = [i for i, v in enumerate(vals) if v]
    return max(nonzero) if nonzero else len(vals) - 1


def py_comparable_ytd(report: ParsedReport, row, cutoff_idx):
    """Sum PY values only across the same months the current year has data for,
    so a partial current year isn't compared against a full PY year."""
    if row is None:
        return None
    total = 0
    found_any = False
    for m in report.month_cols[: cutoff_idx + 1]:
        v = report.month_py(row, m["name"])
        if v is not None:
            found_any = True
            total += v
    return total if found_any else None


def pct_delta(current, prior):
    if not prior:
        return None
    return (current - prior) / prior * 100


# =============================================================================
# SIDEBAR — UPLOADERS
# =============================================================================
st.sidebar.image("https://img.icons8.com/color/96/combo-chart--v1.png", width=60)
st.sidebar.header("QuickBooks Reports Manager")
st.sidebar.markdown(
    "Upload your official QBO exports:\n"
    "- **Profit and Loss by Month** (with prior-year comparison columns enabled)\n"
    "- **Balance Sheet**"
)

pnl_month_file = st.sidebar.file_uploader(
    "1. Profit & Loss by Month Report (.xlsx)", type=["xlsx", "xls"], key="pnl_m"
)
bs_file = st.sidebar.file_uploader(
    "2. Balance Sheet Report (.xlsx)", type=["xlsx", "xls"], key="bs_in"
)

st.sidebar.divider()
process_btn = st.sidebar.button(
    "🚀 Process Financial Dashboard", type="primary", use_container_width=True
)

if not pnl_month_file or not bs_file:
    st.info(
        "👋 **Welcome CFO!** Please upload your **Profit & Loss by Month** and"
        " **Balance Sheet** Excel exports in the sidebar, then click **Process"
        " Financial Dashboard**."
    )
    st.stop()

if not process_btn:
    st.warning(
        "⚠️ Files uploaded. Click **🚀 Process Financial Dashboard** in the"
        " sidebar to generate the executive reports."
    )
    st.stop()

# =============================================================================
# PARSE
# =============================================================================
try:
    pnl = parse_qbo_report(pnl_month_file)
except Exception as e:
    st.error(f"Couldn't parse the P&L by Month file: {e}")
    st.stop()

try:
    bs = parse_qbo_report(bs_file)
except Exception as e:
    st.error(f"Couldn't parse the Balance Sheet file: {e}")
    st.stop()

income_row = pnl.total_of("Income")
expenses_row = pnl.total_of("Expenses")
noi_row = pnl.net_row("Net Operating Income")
net_income_row = pnl.net_row("Net Income")

ytd_revenue = pnl.current_total(income_row) or 0
total_operating_expenses = pnl.current_total(expenses_row) or 0
net_operating_income = pnl.current_total(noi_row) or (ytd_revenue - total_operating_expenses)
net_income = pnl.current_total(net_income_row) if net_income_row else net_operating_income

cutoff_idx = ytd_cutoff_index(pnl, income_row)
ytd_label = pnl.month_cols[cutoff_idx]["name"] if pnl.month_cols else "YTD"

py_revenue = py_comparable_ytd(pnl, income_row, cutoff_idx)
py_opex = py_comparable_ytd(pnl, expenses_row, cutoff_idx)
py_noi = py_comparable_ytd(pnl, noi_row, cutoff_idx)
py_ni = py_comparable_ytd(pnl, net_income_row, cutoff_idx)

# =============================================================================
# EXECUTIVE KPI CARDS
# =============================================================================
st.subheader(f"📌 Executive Performance KPIs (YTD through {ytd_label} vs. Prior Year, same period)")
col1, col2, col3, col4 = st.columns(4)

kpis = [
    (col1, "Total Revenue (YTD)", ytd_revenue, py_revenue),
    (col2, "Operating Expenses", total_operating_expenses, py_opex),
    (col3, "Net Operating Income", net_operating_income, py_noi),
    (col4, "Net Income", net_income, py_ni),
]
for col, label, value, prior in kpis:
    delta = pct_delta(value, prior)
    with col:
        st.metric(
            label=label,
            value=f"${value:,.2f}",
            delta=f"{delta:+.1f}% vs PY" if delta is not None else None,
        )

st.divider()

# =============================================================================
# SECTION 1: MONTH-ON-MONTH EXPENSE TREND
# =============================================================================
st.subheader("📅 Month-on-Month Major Expense Trends")

expense_cats = category_breakdown(pnl, "Expenses")

if not expense_cats:
    st.info("No expense line items were detected under an 'Expenses' section in the uploaded file.")
else:
    active_months = [m["name"] for m in pnl.month_cols[: cutoff_idx + 1]]

    # rebuild per-category monthly series using the same total-row-preferred logic
    indent1 = [r for r in pnl.rows if r["section"] == "Expenses" and r["indent"] == 1]
    total_rows_by_key = {
        r["label"][len("total for "):].strip().lower(): r for r in indent1 if r["is_total"]
    }
    row_by_display_name = {}
    for r in indent1:
        if r["is_total"]:
            row_by_display_name[r["label"][len("Total for "):]] = r
        elif r["label"].strip().lower() not in total_rows_by_key:
            row_by_display_name[r["label"]] = r

    table_rows = []
    for name, _ in expense_cats:
        r = row_by_display_name.get(name)
        row_data = {"Expense Category": name}
        for m in active_months:
            row_data[m] = pnl.month_current(r, m) or 0 if r else 0
        row_data["YTD Total"] = pnl.current_total(r) or 0 if r else 0
        table_rows.append(row_data)

    trend_df = pd.DataFrame(table_rows)
    display_df = trend_df.copy()
    for c in display_df.columns[1:]:
        display_df[c] = display_df[c].apply(lambda x: f"${x:,.2f}")
    st.dataframe(display_df, use_container_width=True, hide_index=True)

st.divider()

# =============================================================================
# SECTION 2: MAJOR COST DRIVERS (YTD)
# =============================================================================
st.subheader("📉 YTD Major Cost Drivers Overview")

if expense_cats:
    drivers_df = pd.DataFrame(expense_cats, columns=["Expense Category", "YTD Total"]).head(10)
    st.bar_chart(drivers_df.set_index("Expense Category"))
    d = drivers_df.copy()
    d["YTD Total"] = d["YTD Total"].apply(lambda x: f"${x:,.2f}")
    st.dataframe(d, use_container_width=True, hide_index=True)

st.divider()

# =============================================================================
# SECTION 3: BALANCE SHEET POSITION
# =============================================================================
st.subheader("🏦 Balance Sheet Position")

total_assets = bs.current_total(bs.total_of("Assets"))
total_liabilities = bs.current_total(bs.total_of("Liabilities"))
total_equity = bs.current_total(bs.total_of("Equity"))
current_assets = bs.current_total(bs.total_of("Current Assets"))
current_liabilities = bs.current_total(bs.total_of("Current Liabilities"))
cash_position = bs.current_total(bs.total_of("Bank Accounts"))
accounts_receivable = bs.current_total(bs.total_of("Accounts Receivable"))

bcol1, bcol2, bcol3, bcol4 = st.columns(4)
with bcol1:
    st.metric("Total Assets", f"${total_assets:,.2f}" if total_assets is not None else "N/A")
with bcol2:
    st.metric(
        "Total Liabilities", f"${total_liabilities:,.2f}" if total_liabilities is not None else "N/A"
    )
with bcol3:
    st.metric("Total Equity", f"${total_equity:,.2f}" if total_equity is not None else "N/A")
with bcol4:
    st.metric("Cash Position", f"${cash_position:,.2f}" if cash_position is not None else "N/A")

bcol5, bcol6, bcol7 = st.columns(3)
with bcol5:
    st.metric(
        "Accounts Receivable", f"${accounts_receivable:,.2f}" if accounts_receivable is not None else "N/A"
    )
with bcol6:
    if current_assets is not None and current_liabilities is not None:
        working_capital = current_assets - current_liabilities
        st.metric("Working Capital", f"${working_capital:,.2f}")
    else:
        st.metric("Working Capital", "N/A")
with bcol7:
    if current_assets and current_liabilities:
        st.metric("Current Ratio", f"{current_assets / current_liabilities:.2f}x")
    else:
        st.metric("Current Ratio", "N/A")

st.divider()

# =============================================================================
# DEBUG / VERIFICATION VIEWS
# =============================================================================
with st.expander("📋 View Parsed P&L Rows (verification)"):
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "Label": r["label"],
                    "Section": r["section"],
                    "Indent": r["indent"],
                    "Is Total": r["is_total"],
                    "YTD Current": pnl.current_total(r),
                    "Full-Year PY": pnl.py_total_full_year(r),
                }
                for r in pnl.rows
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )

with st.expander("📋 View Parsed Balance Sheet Rows (verification)"):
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "Label": r["label"],
                    "Section": r["section"],
                    "Indent": r["indent"],
                    "Is Total": r["is_total"],
                    "Value": bs.current_total(r),
                }
                for r in bs.rows
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
